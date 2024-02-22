from collections import deque
import read_txt as data
import openpyxl

"""
days
month
date
people_num
input_lists:[[name, "〇", "△", "×"...] [...]]
need_people
"""
class person:
    def __init__(self, name, answer, days):
        self.name = name
        self.answer = answer
        self.limit = 4
        self.state = True
        self.workday = 0
        self.shift = [False for i in range(days)]
    def work(self, day):
        if self.workday <= self.limit:
            self.workday += 1
            self.shift[day] = True
        else:
            self.state = False
 
days, month, date, people_num = len(data.days), data.month, data.days, data.people_num
need_people = data.need_people
input_lists = data.input_lists

people = []
name_lists = []
for p in range(people_num):
    input_list = deque(input_lists[p])
    name = input_list.popleft()
    people.append(person(name, input_list, days))
    name_lists.append(name)

def make_shift(people, day):
    need_num = need_people[day]
    num_of_ok = 0
    num_of_soso = 0
    for person in people:
        if person.answer[day] == "◯":
            num_of_ok += 1
        elif person.answer[day] == "△":
            num_of_soso += 1
    if num_of_ok >= need_num:
        while need_num > 0:
            for i, person in enumerate(people):
                if person.answer[day] == "◯":
                    person.work(day)
                    if person.state:
                        need_num -= 1
                if need_num == 0:
                    break
        return
    elif num_of_ok + num_of_soso >= need_num:
        while need_num > 0:
            for i, person in enumerate(people):
                if person.answer[day] == "◯":
                    person.work(day)
                    if person.state:
                        need_num -= 1
            
            for i, person in enumerate(people):
                if person.answer[day] == "△":
                    person.work(day)
                    if person.state:
                        need_num -= 1
                
                if need_num == 0:
                    break
        return
    else:
        for i, person in enumerate(people):
            if person.answer[day] == "◯":
                person.work(day)
                if person.state:
                    need_num -= 1
        
        for i, person in enumerate(people):
            if person.answer[day] == "△":
                person.work(day)
                if person.state:
                    need_num -= 1
        
        for i, person in enumerate(people):
            if person.name == "稲垣" and person.shift[day] == False:
                person.work(day)
            if person.state:
                need_num -= 1
            else:
                person.limit += 1
                person.work(day)
                need_num -= 1
            if need_num == 0:
                break
        else:
            for i, person in enumerate(people):
                if person.name == "小村" and person.shift[day] == False:
                    person.work(day)
                if person.state:
                    need_num -= 1
                else:
                    person.limit += 1
                    person.work(day)
                    need_num -= 1
                if need_num == 0:
                    break
            else:
                print("For day {0} {1} more people are needed".format(day, need_num))
for day in range(days):
    make_shift(people, day)
    people = sorted(people, key=lambda person: person.workday)

can_work_days = [[person.name, []] for person in people]
for k, person in enumerate(people):
    if person.workday == 0 and ("◯" in person.answer or "△" in person.answer):
        for i,ans in enumerate(person.answer):
            if ans == "◯" or ans == "△":
                can_work_days[k][1].append([i])
for i, person in enumerate(people):
    if can_work_days[i][1]:
        [d] = can_work_days[i][1].pop()
        people = sorted(people, key=lambda person: person.workday, reverse = True)
        for one in people:
            if one.shift[d] == True:
                one.shift[d] = False
                person.shift[d] = True
                break
peoples = []
for i in range(people_num):
    for j in range(people_num):
        if name_lists[i] == people[j].name:
            peoples.append(people[j])
            break
people = peoples

wb = openpyxl.Workbook()
sheet = wb.worksheets[0]
for i in range(people_num):
    sheet.cell(1, i+2, people[i].name)
    for j in range(days):
        if people[i].shift[j] == True:
            sheet.cell(j+2, i+2, "〇")
        else:
            sheet.cell(j+2, i+2, "×")
sheet.cell(1,1,"日付")
for i in range(len(date)):
    k = "{0}月{1}日".format(month, date[i])
    sheet.cell(2+i,1, k)
sheet.cell(len(date)+2,1, "表の見方")
sheet.cell(len(date)+3,1,"〇")
sheet.cell(len(date)+3,2,"×")
sheet.cell(len(date)+4, 1, "シフトあり")
sheet.cell(len(date)+4, 2, "シフトなし")

excel_path = r"学習ボランティア{0}月シフト.xlsx".format(month)
wb.save(excel_path)
wb.close()


